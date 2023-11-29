import pymysql
pymysql.install_as_MySQLdb()
import MySQLdb
import xlrd
from openpyxl import load_workbook

database = MySQLdb.connect(host="localhost", user="root", passwd="", db="rfid")
cursor = database.cursor()

# Assuming you have tables named buildings, areas, floors, detail_locations, categories, and items
def get_id_by_name(table_name, name):
    query = f"SELECT id FROM {table_name} WHERE name = %s"
    cursor.execute(query, (name,))
    result = cursor.fetchone()
    return result[0] if result else None

excel_workbook = load_workbook('demo.xlsx')

insert_query = ("INSERT INTO inventories (building_id, area_id, floor_id, detail_location_id, category_id, item_id) VALUES (%s, %s, %s, %s, %s, %s)")

sheet_names = excel_workbook.get_sheet_names()

for sh in range(0, len(sheet_names)):
    # Use get_sheet_by_name() to get the sheet by name
    sheet = excel_workbook[sheet_names[sh]]

    for r in range(2, sheet.max_row + 1):  # Starting from row 2 to skip header
        building_name = sheet.cell(row=r, column=1).value
        area_name = sheet.cell(row=r, column=2).value
        floor_name = sheet.cell(row=r, column=3).value
        detail_location_name = sheet.cell(row=r, column=4).value
        category_name = sheet.cell(row=r, column=7).value
        item_name = sheet.cell(row=r, column=8).value
        quantity = sheet.cell(row=r, column=9).value

        if quantity is None:
            quantity = 0

        quantity_go = quantity + 2

        for h in range(quantity_go):

            # Get the IDs using the get_id_by_name function
            building_id = get_id_by_name("buildings", building_name)
            area_id = get_id_by_name("areas", area_name)
            floor_id = get_id_by_name("floors", floor_name)
            detail_location_id = get_id_by_name("detail_locations", detail_location_name)
            category_id = get_id_by_name("categories", category_name)
            item_id = get_id_by_name("items", item_name)

            product_details_value = (building_id, area_id, floor_id, detail_location_id, category_id, item_id)

            cursor.execute(insert_query, product_details_value)
            database.commit()
        
